from datetime import datetime
from airflow import DAG
from airflow.operators.dummy_operator import DummyOperator
from airflow.operators.python_operator import PythonOperator
from airflow.operators.bash_operator import BashOperator
from airflow.operators.docker_operator import DockerOperator
from docker.types import Mount
import pyodbc
import os
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import pandas as pd
import openpyxl
from connect import *
from pytz import timezone


conn_str_sqlalchemy = (
        f"mssql+pyodbc://{username}:{password}@{IPAddress}/{database_staging}?driver={driver}"
    )
engine_staging = create_engine(conn_str_sqlalchemy, echo=True)

conn_str_sqlalchemy = (
        f"mssql+pyodbc://{username}:{password}@{IPAddress}/{database_source_USA}?driver={driver}"
    )
engine_source_USA = create_engine(conn_str_sqlalchemy, echo=True)

def write_CSV(engine_staging, filename, filepath):
    df = pd.read_csv(filepath)
    df.to_sql(filename, con=engine_staging, if_exists="replace", index=False)
def write_XLSX(engine_staging, filename, filepath):
    df = pd.read_excel(filepath)
    df.to_sql(filename, con=engine_staging, if_exists='replace', index=False)

# procesing Australia

def process_Autralia(engine_staging):
    path_autralia = '/opt/airflow/dags/save/Source/Australia_csv'
    for root, dirs, files in os.walk(path_autralia):
        for file in files:
            print(file)
            if file.endswith(".csv"):
                filename = os.path.splitext(file)[0] 
                filepath = os.path.join(root, file)
                write_CSV(engine_staging,filename, filepath)

# processing Europe
def process_Europe(engine_staging):
    path_Europe = '/opt/airflow/dags/save/Source/Europe_excel/Europe_Sales.xlsx'
    workbook = openpyxl.load_workbook(path_Europe)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        new_workbook = openpyxl.Workbook()
        new_workbook.active.title = sheet_name

        new_sheet = new_workbook.active

        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        output_file = f"/opt/airflow/dags/save/Source/Europe_excel/{sheet_name}.xlsx"
        new_workbook.save(output_file)
        new_workbook.close()

    workbook.close()

    path_USA = '/opt/airflow/dags/save/Source/Europe_excel'
    for root, dirs, files in os.walk(path_USA):
        for file in files:
            if file.endswith(".xlsx") and os.path.splitext(file)[0] != "Europe_Sales":
                filename = os.path.splitext(file)[0]
                filepath = os.path.join(root, file)
                write_XLSX(engine_staging,filename, filepath)

# process common
def process_common(engine_staging):
    path_common = '/opt/airflow/dags/save/Source/Common'
    for root, dirs, files in os.walk(path_common):
        for file in files:
            if file.endswith(".csv"):
                filename = os.path.splitext(file)[0]
                filepath = os.path.join(root, file)
                write_CSV(engine_staging,filename, filepath)
            if file.endswith(".xlsx"):
                filename = os.path.splitext(file)[0]
                filepath = os.path.join(root, file)
                write_XLSX(engine_staging,filename, filepath)

                
def process_USA(engine_source_USA, engine_staging):
    query = '''SELECT * FROM projectBI_Source_USA.dbo.USA_Customer'''
    df = pd.read_sql(query, engine_source_USA)
    df.to_sql('USA_Customer', engine_staging, schema='dbo', index=False, if_exists='replace')
    query = '''SELECT * FROM projectBI_Source_USA.dbo.USA_Person'''
    df = pd.read_sql(query, engine_source_USA)
    df.to_sql('USA_Person', engine_staging, schema='dbo', index=False, if_exists='replace')
    query = '''SELECT * FROM projectBI_Source_USA.dbo.USA_Product'''
    df = pd.read_sql(query, engine_source_USA)
    df.to_sql('USA_Product', engine_staging, schema='dbo', index=False, if_exists='replace')
    query = '''SELECT * FROM projectBI_Source_USA.dbo.USA_SalesOrderDetail'''
    df = pd.read_sql(query, engine_source_USA)
    df.to_sql('USA_SalesOrderDetail', engine_staging, schema='dbo', index=False, if_exists='replace')
    query = '''SELECT * FROM projectBI_Source_USA.dbo.USA_SalesOrderHeader'''
    df = pd.read_sql(query, engine_source_USA)
    df.to_sql('USA_SalesOrderHeader', engine_staging, schema='dbo', index=False, if_exists='replace')
    query = '''SELECT * FROM projectBI_Source_USA.dbo.USA_Staff'''
    df = pd.read_sql(query, engine_source_USA)
    df.to_sql('USA_Staff', engine_staging, schema='dbo', index=False, if_exists='replace')
    query = '''SELECT * FROM projectBI_Source_USA.dbo.USA_Store'''
    df = pd.read_sql(query, engine_source_USA)
    df.to_sql('USA_Store', engine_staging, schema='dbo', index=False, if_exists='replace')
                
default_args = {
    'owner': 'Group20',
    'start_date': datetime(2023, 12, 31, tzinfo=timezone('Asia/Ho_Chi_Minh')),
    'email':'projectBI@gmail.com', 
    'email_on_failure':True
}

with DAG('pipeline_1', default_args=default_args, catchup=True, schedule_interval='00 10 * * *') as dag:
    START = DummyOperator(task_id='START')
    FINISH = DummyOperator(task_id='FINISH_1')
 
    process_Autralia = PythonOperator(task_id='process_Autralia', 
                                    python_callable=process_Autralia, 
                                    op_args = [engine_staging],
                                    dag=dag)
    process_Europe = PythonOperator(task_id='process_Europe', 
                                    python_callable=process_Europe, 
                                    op_args = [engine_staging],
                                    dag=dag)
    process_common = PythonOperator(task_id='process_common', 
                                    python_callable=process_common, 
                                    op_args = [engine_staging],
                                    dag=dag)
    process_USA = PythonOperator(task_id='process_USA', 
                                    python_callable=process_USA, 
                                    op_args = [engine_source_USA, engine_staging],
                                    dag=dag)

    START >> process_Autralia>> process_Europe>> process_common>> process_USA >> FINISH
